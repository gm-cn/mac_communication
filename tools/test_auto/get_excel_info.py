# -*- coding:utf-8 -*-
import xlrd
import os
import xlrd
import copy
import shutil
from openpyxl import load_workbook
from configparser import ConfigParser
from database import Database_test

cp = ConfigParser()
cp.read("bms.ini")

EXCEL_PATH = cp.get("excel", "path")


class Excel(object):
    def __init__(self):
        self.filepath = cp.get("excel", "path")
        self.nic_count = int(cp.get('network', 'count'))

    def open_excel(self):
        excel = xlrd.open_workbook(self.filepath, encoding_override="utf-8")
        sheets = excel.sheet_names()
        for i in sheets:
            table = excel.sheet_by_name(i)
            header_row = table.row_values(0)
            if u"idrac地址" in header_row:
                return table

    def get_userinfo(self):
        table = self.open_excel()
        header_row = table.row_values(0)
        datainfo = table.row_values(1)
        ip_heads = [u'交换机用户名', u'交换机密码', u'带外账号', u'带外密码', u'BootMode']
        res = {}
        for i in ip_heads:
            port_index = header_row.index(i)
            data_port = datainfo[port_index]
            res[i] = data_port
        return res

    def get_ips(self):
        table = self.open_excel()
        header_row = table.row_values(0)

        ip_heads = [u'idrac地址']
        for i in range(1, self.nic_count + 1):
            ip_heads.append(u'端口%sip' % i)
        ipmi_ips = {}
        for i in ip_heads:
            port_index = header_row.index(i)
            data_port = table.col_values(port_index)
            ipmi_ips[i] = data_port[1:]
        all_ips = []
        for i in range(table.nrows - 1):
            ips = []
            for j in ip_heads:
                ips.append(ipmi_ips[j][i])
            all_ips.append(ips)
        return all_ips

    def get_pxe_port(self):
        table = self.open_excel()
        header_row = table.row_values(0)
        port_index = header_row.index(u'端口1')
        switch_index = header_row.index(u'交换机1地址')
        res = {}
        data_port = table.col_values(port_index)
        data_ip = table.col_values(switch_index)
        data = [[data_port[i], data_ip[i]] for i in range(1, len(data_port))]

        for i in set(data_ip[1:]):
            res[i] = []
        for i in data:
            res[i[1]].append(i[0])
        return res

    def get_other_port(self):
        table = self.open_excel()
        header_row = table.row_values(0)
        res = {}
        for i in range(2, self.nic_count + 1):
            port_index = header_row.index(u'端口%s' % i)
            switch_index = header_row.index(u'交换机%s地址' % i)
            data_port = table.col_values(port_index)
            data_ip = table.col_values(switch_index)
            data = [[data_port[i], data_ip[i]] for i in range(1, len(data_port))]

            for i in set(data_ip[1:]):
                res[i] = []
            for i in data:
                res[i[1]].append(i[0])
        return res

    def get_other_suc_port(self, now_time):
        table = self.open_excel()
        header_row = table.row_values(0)
        res = {}

        ipmi_index = header_row.index(u"idrac地址")
        ipmi_ips = table.col_values(ipmi_index)
        now_index = [ipmi_ips.index(i) for i in now_time]

        for i in range(2, self.nic_count + 1):
            port_index = header_row.index(u'端口%s' % i)
            switch_index = header_row.index(u'交换机%s地址' % i)
            data_port_all = table.col_values(port_index)
            data_ip_all = table.col_values(switch_index)
            data_port = [data_port_all[dp] for dp in now_index]
            data_ip = [data_ip_all[di] for di in now_index]
            data = [[data_port[i], data_ip[i]] for i in range(len(data_port))]

            for i in set(data_ip):
                res[i] = []
            for i in data:
                res[i[1]].append(i[0])
        return res

    def get_all_port(self):
        table = self.open_excel()
        header_row = table.row_values(0)
        res = {}
        for i in range(1, self.nic_count + 1):
            port_index = header_row.index(u'端口%s' % i)
            switch_index = header_row.index(u'交换机%s地址' % i)
            data_port = table.col_values(port_index)
            data_ip = table.col_values(switch_index)
            data = [[data_port[i], data_ip[i]] for i in range(1, len(data_port))]

            for i in set(data_ip[1:]):
                res[i] = []
            for i in data:
                res[i[1]].append(i[0])
        return res

    def get_all_switch(self):
        table = self.open_excel()
        header_row = table.row_values(0)
        sw_ips = []
        for i in range(1, self.nic_count + 1):
            switch_index = header_row.index(u'交换机%s地址' % i)
            data_ip = table.col_values(switch_index)
            sw_ips.extend(data_ip[1:])
        return list(set(sw_ips))



def gen_log():
    now_path = os.getcwd()
    pro_path = now_path.split('tools')[0]
    filepath = pro_path + 'auto_test/host_ip_all'
    filepath1 = pro_path + 'auto_test/host_ip'

    if os.path.exists('pxe_log') and os.path.exists('image_log') and os.path.exists(filepath):
        os.remove('pxe_log')
        os.remove('image_log')
        os.remove(filepath)

    with Database_test() as data_t:
        count = data_t.nic_count
        hosts1 = data_t.gen_host_ip()
        hosts2 = data_t.gen_install_success()

    tem1 = [i[0] for i in hosts1]
    pxe_success = [i[0] for i in hosts2]
    pxe_failed = [i for i in tem1 if i not in pxe_success]
    install_success = [i[0] for i in hosts2 if 'success' in i]
    install_failed = [i for i in pxe_success if i not in install_success]
    host_ip = [i for i in hosts1 if i[0] in install_success]
    if os.path.exists('success'):
        with open('success', 'r') as fp:
            last_time = fp.readlines()
    else:
        last_time = []
    last_time = [i.replace('\n', '') for i in last_time]
    now_time = [i for i in install_success if i not in last_time]

    with open('success', 'w') as fp:
        for i in install_success:
            fp.write(i)
            fp.write('\n')

    with open('pxe_log', 'a') as fp:
        for i in pxe_failed:
            fp.write("%s pxe boot error\n" % i)
        for i in pxe_success:
            fp.write("%s pxe boot success\n" % i)

    with open('image_log', 'a') as fp:
        for i in install_failed:
            fp.write("%s install image error\n" % i)

        for i in pxe_failed:
            fp.write("%s install image error\n" % i)

        for i in install_success:
            fp.write("%s install image success\n" % i)

    with open(filepath, 'a') as fp:
        for i in host_ip:
            for j in range(1, count + 1):
                fp.write(i[j])
                if j != count:
                    fp.write(' ')
            fp.write('\n')

    with open(filepath1, 'w') as fp:
        for i in host_ip:
            if i[0] in now_time:
                for j in range(1, count + 1):
                    fp.write(i[j])
                    if j != count:
                        fp.write(' ')
                fp.write('\n')

    ex = Excel()
    sw_info = ex.get_userinfo()
    sw_ip = ex.get_all_switch()
    with open(pro_path + 'auto_test/switch_ip', 'w') as fp:
        for i in sw_ip:
            fp.write("{} {} {}\n".format(i, sw_info[u'交换机用户名'], sw_info[u'交换机密码']))

    return now_time


def checkout_excel():
    with Database_test() as data_t:
        count = data_t.nic_count
        hosts = data_t.gen_install_success()

    ipmi_ips = [i[0] for i in hosts]

    excel = xlrd.open_workbook(EXCEL_PATH, encoding_override="utf-8")
    sheets = excel.sheet_names()
    sheet_name = ''
    for s_name in sheets:
        table = excel.sheet_by_name(s_name)
        header_row = table.row_values(0)
        if u"idrac地址" in header_row:
            sheet_name = s_name
            break

    shutil.copyfile(EXCEL_PATH, 'base_new.xlsx')
    workbooknew = load_workbook('base_new.xlsx')

    table = excel.sheet_by_name(sheet_name)
    sheets_copy = workbooknew[sheet_name]
    rows_num = table.nrows
    header_row = table.row_values(0)

    switch_indexs = []
    mac_indexs = []
    for c in range(1, count + 1):
        mac_index = header_row.index(u'网卡mac%s' % c)
        switch_index = header_row.index(u'交换机%s地址' % c)
        mac_indexs.append(mac_index)
        switch_indexs.append(switch_index)

    for i in range(1, rows_num):
        init_info = table.row_values(i)
        ipmi_ip = init_info[1]
        try:
            index_new = ipmi_ips.index(ipmi_ip)
        except:
            continue

        init_macs = [init_info[mac_ind] for mac_ind in mac_indexs]
        init_switchs = [init_info[sw_ind] for sw_ind in switch_indexs]

        new_macs = [m.upper() for m in hosts[index_new][1:1 + count]]

        new_switchs = []
        mac_res = [m1 for m1 in init_macs if m1 in new_macs]
        if len(mac_res) != len(new_macs):
            print('%s mac address has error' % ipmi_ip)
            for mac_id in range(len(new_macs)):
                sheets_copy.cell(i + 1, mac_indexs[mac_id] + 1).value = new_macs[mac_id]
            with open('excel_log', 'a') as fp:
                fp.write('%s mac address error' % ipmi_ip)
            init_macs = new_macs

        with Database_test() as data_t:
            dhcp_mac = data_t.select_createbms_info('dhcp_mac', ipmi_ip)
            dhcp_mac = dhcp_mac[0]

        new_macs1 = copy.deepcopy(new_macs)
        if new_macs1[0] != dhcp_mac:
            print('%s Port order has error' % ipmi_ip)
            try:
                ind = new_macs1.index(dhcp_mac)
            except:
                continue
            new_macs1[0], new_macs1[ind] = new_macs1[ind], new_macs1[0]
            for id in range(len(new_macs1)):
                ind = init_macs.index(new_macs1[id])
                new_switchs.append(init_switchs[id])
                sheets_copy.cell(i+1, switch_indexs[ind] + 1).value = init_switchs[id]
        else:
            continue

    workbooknew.save('base_new.xlsx')
    return True
